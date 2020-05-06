#!/usr/bin/env python
# enccoding: utf-8
'''
@author: Yankai
@project:
'''
import openpyxl
#打开excel
def readExcel():
    filename = r'C:\Users\kmf\接口复习\test.xlsx'
    inwb = openpyxl.load_workbook(filename)
    #获取打开的excel的sheet内容
    sheetnames = inwb.get_sheet_names()#获取读文件中所有的sheet，通过名字的方式
    print(sheetnames)
    ws = inwb.get_sheet_by_name(sheetnames[0])#获取第一个sheet内容
    print(ws)
    #获取sheet的最大行数和列数
    rows = ws.max_row
    cols =ws.max_column
    print(rows,cols)
#打开将写的表并添加sheet
def writeExcel():
    outwb = openpyxl.load_workbook('C:/Users\kmf\接口复习/test.xlsx')
    outws = outwb.create_sheet(index=2,title='sheet3')
#保存
    saveExcel = 'C:/Users\kmf\接口复习/test.xlsx'
    outwb.save(saveExcel)

if __name__ == '__main__':
    readExcel()
    writeExcel()